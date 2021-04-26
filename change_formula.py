
import openpyxl
import tkinter.messagebox as msgbox

weekly_path = r'\\server05\GMG Tree\All GMG\Gmg\Weekly Meetings\Meterages\Weekly Meterage 300421.xlsm'
weekly_wb = openpyxl.load_workbook(weekly_path, read_only=False, keep_vba=True)
weekly_ws_render = weekly_wb["Render Squads"]

# change formula in render sheet
def update_formula():
    col_total = get_total_col()
    row_total = get_total_row()
    formula = weekly_ws_render.cell(row_total, col_total).value
    num = formula[len(formula)-3:len(formula)-1]
    num_new = int(num)+1
    # total
    weekly_ws_render.cell(row_total, col_total).value = formula.replace(num, str(num_new))
    # total meter
    formula_meter = weekly_ws_render.cell(row_total, col_total-1).value
    weekly_ws_render.cell(row_total, col_total-1).value = formula_meter.replace(num, str(num_new))
    # total worker
    formula_worker = weekly_ws_render.cell(row_total, col_total - 3).value
    weekly_ws_render.cell(row_total, col_total-3).value = formula_worker.replace(num, str(num_new))

def get_total_col():
    for col in range(weekly_ws_render.max_column, 10, -1):
        if weekly_ws_render.cell(1, col).value == 'Total':
            return col
            break

def get_total_row():
    for row in range(weekly_ws_render.max_row, 20, -1):
        if weekly_ws_render.cell(row, 3).value == "Total Meterage/Days":
            return row
            break


update_formula()
#msgbox.showinfo(message="done")
weekly_wb.save(r'\\server05\GMG Tree\All GMG\Gmg\Weekly Meetings\Meterages\Weekly Meterage 300421.xlsm')
print("done!")
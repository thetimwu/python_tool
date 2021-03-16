import openpyxl

weekly_path = 'C:/P/gmg/gmg_spreadsheet/Weekly Meterage 190321.xlsm'
weekly_wb = openpyxl.load_workbook(weekly_path, read_only=True)
weekly_ws_tiling = weekly_wb["Tiling"]
weekly_ws_render = weekly_wb["Render"]
weekly_ws_joiner = weekly_wb["Joiner"]

weekly_roofer = []
weekly_tiler = []
weekly_hc = []
weekly_vc = []
weekly_stripping = []
weekly_render = []
weekly_beading = []
weekly_ingoes = []
weekly_cills = []
weekly_joiner = []
weekly_scrimming = []
weekly_fungwash = []


# tiling sheet
def pull_names_from_weekly_meterage_titling():
    for row_weekly in range(2, weekly_ws_tiling.max_row):
        name = weekly_ws_tiling.cell(row_weekly, 1).value
        if not name:
            continue

        if "tiler" in weekly_ws_tiling.cell(row_weekly, 2).value.lower() and not weekly_ws_tiling.cell(row_weekly, 3).value:
            if name not in weekly_tiler:
                weekly_tiler.append(name)
        elif "flat" in weekly_ws_tiling.cell(row_weekly, 2).value.lower() and not weekly_ws_tiling.cell(row_weekly, 3).value:
            if name not in weekly_roofer:
                weekly_roofer.append(name)
        elif "HC" == weekly_ws_tiling.cell(row_weekly, 3).value:
            if name not in weekly_hc:
                weekly_hc.append(name)
        elif "VC" in weekly_ws_tiling.cell(row_weekly, 3).value:
            if name not in weekly_vc:
                weekly_vc.append(name)
        elif "Stripping" in weekly_ws_tiling.cell(row_weekly, 3).value:
            if name not in weekly_stripping:
                weekly_stripping.append(name)


#render sheet
def pull_names_from_weekly_meterage_render():
    for row_weekly in range(2, weekly_ws_render.max_row):
        name = weekly_ws_render.cell(row_weekly, 1).value
        if not name:
            continue

        if "roughcaster" in weekly_ws_render.cell(row_weekly, 2).value.lower() and not weekly_ws_render.cell(row_weekly, 3).value:
            if name not in weekly_render:
                weekly_render.append(name)
        elif "beading" in weekly_ws_render.cell(row_weekly, 2).value.lower():
            if name not in weekly_beading:
                weekly_beading.append(name)
        elif "scrimming" in weekly_ws_render.cell(row_weekly, 2).value.lower():
            if name not in weekly_scrimming:
                weekly_scrimming.append(name)
        elif "ingoes" in weekly_ws_render.cell(row_weekly, 2).value.lower():
            if name not in weekly_ingoes:
                weekly_ingoes.append(name)
        elif "cills" in weekly_ws_render.cell(row_weekly, 2).value.lower():
            if name not in weekly_cills:
                weekly_cills.append(name)
        elif "cills" in weekly_ws_render.cell(row_weekly, 2).value.lower():
            if name not in weekly_cills:
                weekly_cills.append(name)
        elif "fungwash" in weekly_ws_render.cell(row_weekly, 2).value.lower():
            if name not in weekly_fungwash:
                weekly_fungwash.append(name)
        elif weekly_ws_render.cell(row_weekly, 3).value:
            if "beading" in weekly_ws_render.cell(row_weekly, 3).value.lower():
                if name not in weekly_beading:
                    weekly_beading.append(name)


def pull_names_from_weekly_meterage_joiner():
    for row_weekly in range(2, weekly_ws_joiner.max_row):
        name = weekly_ws_joiner.cell(row_weekly, 1).value
        if name not in weekly_joiner:
            weekly_joiner.append(name)

pull_names_from_weekly_meterage_titling()
pull_names_from_weekly_meterage_render()
pull_names_from_weekly_meterage_joiner()
#loop yearly meterage

print(weekly_joiner)
weekly_wb.close()

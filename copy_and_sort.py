# copy worksheet from one to another, delete empty rows in between and sort
import openpyxl as xl

file_name = "Weekly Meterage 160421.xlsm"
weekly_path = 'C:/p/gmg/gmg_spreadsheet/'+file_name
weekly_wb = xl.load_workbook(weekly_path, read_only=False, keep_vba=True, data_only=True)
weekly_ws_render = weekly_wb["Render Squads"]
weekly_ws_tiling = weekly_wb["Tiling Squads"]
ws_asc_render = weekly_wb["Render Squads Ascending"]
ws_asc_tiling = weekly_wb["Tiling Squads Ascending"]

# copy cell values from source wks to destination wks
def copy_range(wks):
    rangeSelected = []
    mr = wks.max_row
    mc = get_max_col(wks)-1
    print("mr: ", mr, "mc: ", mc)
    for i in range(2, mr+1):
        rowSelected = []
        for j in range(1, mc-2):
            rowSelected.append(wks.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)
        #print(rowSelected[len(rowSelected)-6])
    return rangeSelected

def sort_by_total(e):
    # column total
    return e[len(e)-6]

def sort_wks(wks):
    copied_range = copy_range(wks)
    copied_range.sort(reverse=True, key=sort_by_total)
    mr = wks.max_row
    mc = get_max_col(wks) - 1
    for i in range(2, mr+1):
        for j in range(1, mc-2):
            wks.cell(i,j).value = copied_range[i-2][j-1]


def get_max_col(wks):
    for col in range(10, wks.max_column):
        if weekly_ws_render.cell(1, col).value is None:
            return col
            break


def get_max_row(wks, init=20):
    for row in range(init, wks.max_row):
        if wks.cell(row, 2).value is None:
            return row
            break


def get_none_empty_row(wks, init):
    for row in range(init, wks.max_row):
        if not wks.cell(row, 2).value is None:
            return row
            break


def delete_empty_row(wks):
    mr1 = get_max_row(wks)
    none_empty_row = get_none_empty_row(wks, mr1)
    # for titling squads sheet only
    if wks == weekly_wb["Tiling Squads Ascending"]:
        none_empty_row = get_none_empty_row(wks, none_empty_row+3)
    mr2 = get_max_row(wks, none_empty_row)
    #print("mr1:", mr1, ", none_empty_row: ", none_empty_row, "mr2: ", mr2)
    # delete the rest rows of the spreadsheet
    wks.delete_rows(mr2, 100)
    # delete empty rows in between
    wks.delete_rows(mr1, none_empty_row - mr1)

# def sort_sheet_by_total(wks):
#     wks.

weekly_wb.remove(ws_asc_tiling)
asc_tiling = weekly_wb.copy_worksheet(weekly_ws_tiling)
asc_tiling.title = "Tiling Squads Ascending"
mc_tiling = get_max_col(asc_tiling)
asc_tiling.delete_cols(mc_tiling-3, 3)
delete_empty_row(asc_tiling)
#asc_tiling.sort()
sort_wks(asc_tiling)

weekly_wb.remove(ws_asc_render)
asc_render = weekly_wb.copy_worksheet(weekly_ws_render)
asc_render.title = "Render Squads Ascending"
mc_render = get_max_col(asc_render)
asc_render.delete_cols(mc_render-3, 3)
delete_empty_row(asc_render)
sort_wks(asc_render)


#msgbox.showinfo(message="done")
weekly_wb.save('C:/P/gmg/gmg_spreadsheet/'+file_name)
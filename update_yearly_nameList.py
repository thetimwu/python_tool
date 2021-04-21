import openpyxl as xl
from openpyxl.styles.borders import Border, Side
import tkinter.messagebox as msgbox
# update yearly name list by comparing name list from weekly wks.

file_name = "Weekly Meterage 160421.xlsm"
file_name_yearly = "C:/p/gmg/gmg_spreadsheet/Jan-Dec2021.xlsm"
weekly_path = 'C:/p/gmg/gmg_spreadsheet/'+file_name
weekly_wb = xl.load_workbook(weekly_path, read_only=False, keep_vba=True)
yearly_wb = xl.load_workbook(file_name_yearly, read_only=False, keep_vba=True)
weekly_ws_render = weekly_wb["Render"]
weekly_ws_tiling = weekly_wb["Tiling"]
weekly_ws_joiner = weekly_wb["Joiner"]

# copy cell values from source wks to destination wks
name_tiling_weekly = []
name_render_weekly = []
name_beading_weekly = []
name_scrimming_weekly = []
name_ingoes_weekly = []
name_lath_weekly = []
name_hc_weekly = []
name_vc_weekly = []
name_stripping_weekly = []
name_cills_weekly = []
name_joiner_weekly = []
name_fungwash_weekly = []

name_tiling_yearly = []
name_render_yearly = []
name_beading_yearly = []
name_scrimming_yearly = []
name_ingoes_yearly = []
name_lath_yearly = []
name_hc_yearly = []
name_vc_yearly = []
name_stripping_yearly = []
name_cills_yearly = []
name_joiner_yearly = []
name_fungwash_yearly = []

def populate_list(wks, criteria):
    name_list = []
    mr = wks.max_row
    for i in range(2, mr):
        na = wks.cell(i, 1).value
        if na is not None:
            if wks.cell(i, 3).value == criteria and na not in name_list:
                name_list.append(na)
            elif wks.cell(i, 2).value == criteria and na not in name_list and wks.cell(i, 3).value is None:
                name_list.append(na)
            elif criteria == "tiling" and na not in name_list and wks.cell(i, 3).value is None:
                name_list.append(na)
            elif criteria == "Roughcaster" and wks.cell(i, 2).value == "Roughcaster" and na not in name_list and wks.cell(i, 3).value is None:
                name_list.append(na)

    return name_list

def populate_list_yearly(wks):
    name_list = []
    mr = get_max_row(wks)

    for i in range(2, mr):
        na = wks.cell(i, 1).value
        if na not in name_list and na is not None:
            name_list.append(na)

    return name_list



def get_max_col(wks):
    for col in range(10, wks.max_column):
        if weekly_ws_render.cell(1, col).value is None:
            return col
            break


def get_max_row(wks, init=2):
    for row in range(init, wks.max_row):
        if wks.cell(row, 1).value is None:
            return row
            break


def get_none_empty_row(wks, init):
    for row in range(init, wks.max_row):
        if not wks.cell(row, 2).value is None:
            return row
            break

name_tiling_weekly = populate_list(weekly_ws_tiling, "tiling")
name_render_weekly = populate_list(weekly_ws_render, "Roughcaster")
name_beading_weekly = populate_list(weekly_ws_render, "Beading")
name_scrimming_weekly = populate_list(weekly_ws_render, "Scrimming")
name_ingoes_weekly = populate_list(weekly_ws_render, "Ingoes")
name_lath_weekly = populate_list(weekly_ws_render, "Lath")
name_hc_weekly = populate_list(weekly_ws_tiling, "HC")
name_vc_weekly = populate_list(weekly_ws_tiling, "VC")
name_stripping_weekly = populate_list(weekly_ws_tiling, "Stripping")
name_cills_weekly = populate_list(weekly_ws_render, "Cills")
name_joiner_weekly = populate_list(weekly_ws_joiner, "Joiner")
name_fungwash_weekly = populate_list(weekly_ws_render, "FungWash")

name_tiling_yearly = populate_list_yearly(yearly_wb["Tiling"])
name_render_yearly = populate_list_yearly(yearly_wb["Render"])
name_beading_yearly = populate_list_yearly(yearly_wb["Beading"])
name_scrimming_yearly = populate_list_yearly(yearly_wb["Scrimming"])
name_ingoes_yearly = populate_list_yearly(yearly_wb["Ingoes"])
#name_lath_yearly = populate_list_yearly(yearly_wb[""])
name_hc_yearly = populate_list_yearly(yearly_wb["HC"])
name_vc_yearly = populate_list_yearly(yearly_wb["VC"])
name_stripping_yearly = populate_list_yearly(yearly_wb["Stripping"])
name_cills_yearly = populate_list_yearly(yearly_wb["Cills"])
name_joiner_yearly = populate_list_yearly(yearly_wb["Joiner"])
name_fungwash_yearly = populate_list_yearly(yearly_wb["FungWash"])

def update_yearly(wks, list_weekly, list_yearly):
    mr = get_max_row(wks)
    for i in list_weekly:
        if i not in list_yearly:
            print("found: ", i)
            wks.insert_rows(mr)
            wks.cell(mr, 1).value = i
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            # add border to the inserted row
            for j in range(1, 16):
                wks.cell(mr,j).border = thin_border



# print("weekly fgwsh:", name_fungwash_weekly)
# print("yearly fgwsh: ", name_fungwash_yearly)

update_yearly(yearly_wb["FungWash"], name_fungwash_weekly, name_fungwash_yearly)

weekly_wb.close()
# yearly_wb.close()
yearly_wb.save(file_name_yearly)